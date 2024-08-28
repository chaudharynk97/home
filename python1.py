from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

# Step 1: Load the original workbook and select the worksheet
wb_original = load_workbook("file_example_XLSX_1000.xlsx")
ws_original = wb_original["Sheet1"]

# Load the second workbook with values in column G
wb_values = load_workbook("Book1.xlsx")
ws_values = wb_values["Sheet1"]

# Create a new workbook and get the active sheet
new_wb = Workbook()
new_ws = new_wb.active

# Step 2: Copy headers from the original workbook and append to column A
headers = [cell.value for cell in ws_original[1]]
new_ws.append(headers)

# Step 3: Read values from column G of the second workbook (skipping the header)
values_in_g = [cell.value for cell in ws_values['G']]

# Step 4: Track found and not found values
found_values = set()
not_found_values = []

# Step 5: Iterate over the original worksheet and find matching rows
for row in ws_original.iter_rows(min_row=2):  # Skipping the header row
    if row[6].value in values_in_g:  # Column G is index 6 (0-based index)
        found_values.add(row[6].value)
        # Get data from column G onwards
        row_data = [None] * 6 + [cell.value for cell in row[6:]]
        new_ws.append(row_data)

# Step 6: Track values that were not found
for value in values_in_g:
    if value not in found_values:
        not_found_values.append(value)

# Step 7: Append not found values with "not found" message in red, starting from column G
for value in not_found_values:
    new_ws.append([None] * 6 + [value, "not found"])  # Insert "None" for columns A to F
    cell = new_ws.cell(row=new_ws.max_row, column=8)  # Get the cell with "not found" (H column)
    cell.font = Font(color="FF0000")  # Set font color to red

# Step 8: Save the new workbook
new_wb.save("filtered_example.xlsx")

print("Filtered data saved to filtered_example.xlsx")
